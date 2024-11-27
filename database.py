import os
from statistics import mode
import openpyxl
import cv2
import numpy as np
import torch.utils.data.dataset
import utils.functional as T
from torch.utils.data import DataLoader
import torchvision.transforms as transforms
from functools import partial


def parse_images(data_root, mode):
    excel_dir = os.path.join(data_root, 'excel', mode)
    excel_list = os.listdir(excel_dir)
    image_info = []
    new_k = []
    for file in excel_list:
        excel = openpyxl.load_workbook(os.path.join(data_root, 'excel', mode, file))
        sheet = excel['data']
        for row in sheet.values:
            poc, ctu_addr, C, K, R1, R2, R3, R4, R5, D1, D2, D3, D4, D5 = row
            if C <= 0.8 or C >= 3200:
                continue
            if K <= -3 or K >= 0:
                continue
            para = (C, K)
            # para = K
            new_k.append(para)
            sample = (R1, R2, R3, R4, R5, D1, D2, D3, D4, D5)
            box = ((ctu_addr % 2) * 128, (ctu_addr // 2) * 128, 128, 128)
            image_dir = os.path.join(data_root, mode, file.split('.')[0], '%04d.png' % (poc+1))
            item = (image_dir, box, para, sample)

            image_info.append(item)
    print('Image Parse Completed')
    return image_info


class IntraDataset(torch.utils.data.Dataset):
    def __init__(
        self,
        data_root,
        mode,
    ):
        self.data_root = data_root
        self.mode = mode
        self.image_info = parse_images(self.data_root, self.mode)
        self.transform_test = transforms.Compose([
            transforms.ToTensor(),
            transforms.Resize((256, 256)),
            # transforms.RandomCrop((min(args.height, args.weight))),
        ])
        self.transform_train = transforms.Compose([
            transforms.ToTensor(),
            transforms.Resize((256, 256)),
            transforms.RandomRotation(30),
            transforms.RandomCrop((224)),
            transforms.RandomHorizontalFlip(),
            transforms.RandomVerticalFlip(),
            transforms.RandomErasing(p=1, scale=(
                0.0002, 0.0005), ratio=(0.3, 3.3)),
            transforms.RandomAdjustSharpness(7, p=0.5),
        ])

    def __getitem__(self, index):
        (
            image_dir,
            box,
            para,
            sample
        ) = self.image_info[index]
        try:
            img = cv2.imread(image_dir)
            img_ = img[box[1]:box[1]+box[3], box[0]:box[0]+box[2], :]
            img_crop = cv2.cvtColor(img_, cv2.COLOR_BGR2YUV)
            # img_crop = T.to_mytensor(img_crop)
            img_crop = img_crop / 255.

            outputs = [
                img_crop,
                para,
                sample
            ]
            if self.mode == 'train':
                img_crop = self.transform_train(img_crop)
            else:
                img_crop = self.transform_test(img_crop)

        except Exception as e:
            print("problem in, ", image_dir)
            print(e)
            return self.__getitem__(np.random.randint(0, len(self.image_info)))
        return outputs

    def __len__(self):
        return len(self.image_info)


def parse_test_images(data_root):
    excel_dir = os.path.join(data_root, 'excel')
    excel_list = os.listdir(excel_dir)
    image_info = []
    for file in excel_list:
        seq_name = file.split('.')[0]
        img_path = os.path.join(data_root, seq_name, '0001.bmp')
        img = cv2.imread(img_path)
        h, w, c = img.shape
        col = w // 128
        excel = openpyxl.load_workbook(os.path.join(data_root, 'excel', file))
        sheet = excel['data']
        for row in sheet.values:
            poc, ctu_addr, C, K, R1, R2, R3, R4, R5, D1, D2, D3, D4, D5 = row
            para = (C, K)
            sample = (R1, R2, R3, R4, R5, D1, D2, D3, D4, D5)
            box = ((ctu_addr % col) * 128, (ctu_addr // col) * 128, 128, 128)
            item = (img_path, box, para, sample)
            image_info.append(item)
    print('Image Parse Completed')
    return image_info


class IntraTestSet(torch.utils.data.Dataset):
    def __init__(
            self,
            data_root,
    ):
        self.data_root = data_root
        self.image_info = parse_test_images(self.data_root)

    def __getitem__(self, index):
        (
            image_dir,
            box,
            para,
            sample
        ) = self.image_info[index]
        try:
            img = cv2.imread(image_dir)
            img_ = img[box[1]:box[1] + box[3], box[0]:box[0] + box[2], :]
            img_crop = cv2.cvtColor(img_, cv2.COLOR_BGR2YUV)
            # img_crop = T.to_mytensor(img_crop)
            img_crop = img_crop / 255.

            outputs = [
                img_crop,
                para,
                sample
            ]

        except Exception as e:
            print("problem in, ", image_dir)
            print(e)
            return self.__getitem__(np.random.randint(0, len(self.image_info)))
        return outputs

    def __len__(self):
        return len(self.image_info)


def parse_frame(data_root, width, height):
    image_info = []
    row = height // 128
    col = width // 128
    for imgs_dir in os.listdir(data_root):
        for i in range(row):
            for j in range(col):
                box = (j * 128, i * 128, 128, 128)
                image_dir = os.path.join(data_root, imgs_dir)
                item = (image_dir, box)
                image_info.append(item)
    return image_info


class FrameDataset(torch.utils.data.Dataset):
    def __init__(
        self,
        data_root,
        width,
        height
    ):
        self.data_root = data_root
        self.mode = mode
        self.height = height
        self.width = width
        self.image_info = parse_frame(self.data_root, self.width, self.height)

    def __getitem__(self, index):
        (
            image_dir,
            box
        ) = self.image_info[index]
        try:
            img = cv2.imread(image_dir)
            img_ = img[box[1]:box[1]+box[3], box[0]:box[0]+box[2], :]
            img_crop = cv2.cvtColor(img_, cv2.COLOR_BGR2YUV)
            # img_crop = T.to_mytensor(img_crop)
            img_crop = img_crop / 255.

            outputs = img_crop

        except Exception as e:
            print("problem in, ", image_dir)
            print(e)
            return self.__getitem__(np.random.randint(0, len(self.image_info)))
        return outputs

    def __len__(self):
        return len(self.image_info)


def parse_seq(data_root, seq_name, width):
    excel_dir = os.path.join(data_root, 'excel', seq_name + '.xlsx')
    excel = openpyxl.load_workbook(excel_dir)
    sheet = excel['data']
    image_info = []
    col = width // 128
    for row in sheet.values:
        poc, ctu_addr, C, K, R1, R2, R3, R4, R5, D1, D2, D3, D4, D5 = row
        if C <= 0.8 or C >= 3200:
            continue
        if K <= -3 or K >= 0:
            continue
        para = (C, K)
        sample = (R1, R2, R3, R4, R5, D1, D2, D3, D4, D5)
        box = ((ctu_addr % col) * 128, (ctu_addr // col) * 128, 128, 128)
        image_dir = os.path.join(data_root, 'TestSet', seq_name, '%04d.bmp' % (poc+1))
        item = (image_dir, box, para, sample)
        image_info.append(item)
        img = cv2.imread(image_dir)
        img_ = img[box[1]:box[1] + box[3], box[0]:box[0] + box[2], :]
        cv2.imwrite('test.png', img_)
    print('Image Parse Completed')
    return image_info


class SeqDataset(torch.utils.data.Dataset):
    def __init__(
        self,
        data_root,
        seq_name,
        width
    ):
        self.data_root = data_root
        self.seq_name = seq_name
        self.width = width
        self.image_info = parse_seq(self.data_root, self.seq_name, self.width)
        self.transform_test = transforms.Compose([
            transforms.ToTensor(),
            transforms.Resize((256, 256)),
            # transforms.RandomCrop((min(args.height, args.weight))),
        ])
        self.transform_train = transforms.Compose([
            transforms.ToTensor(),
            transforms.Resize((256, 256)),
            transforms.RandomRotation(30),
            transforms.RandomCrop((224)),
            transforms.RandomHorizontalFlip(),
            transforms.RandomVerticalFlip(),
            transforms.RandomErasing(p=1, scale=(
                0.0002, 0.0005), ratio=(0.3, 3.3)),
            transforms.RandomAdjustSharpness(7, p=0.5),
        ])

    def __getitem__(self, index):
        (
            image_dir,
            box,
            para,
            sample
        ) = self.image_info[index]
        try:
            img = cv2.imread(image_dir)
            img_ = img[box[1]:box[1]+box[3], box[0]:box[0]+box[2], :]
            img_crop = cv2.cvtColor(img_, cv2.COLOR_BGR2YUV)
            img_crop = img_crop / 255.

            outputs = [
                img_crop,
                para,
                sample
            ]

        except Exception as e:
            print("problem in, ", image_dir)
            print(e)
            return self.__getitem__(np.random.randint(0, len(self.image_info)))
        return outputs

    def __len__(self):
        return len(self.image_info)


def parse_hdr(excel_root, data_root, mode):
    excel_list = os.listdir(os.path.join(excel_root, mode))
    image_info = []
    for file in excel_list:
        excel = openpyxl.load_workbook(os.path.join(excel_root, mode, file))
        sheet = excel['data']
        for row in sheet.values:
            poc, ctu_addr, C, K, theta, C_Y, K_Y, theta_Y, factor_22, factor_27, factor_32, factor_37, factor_42, \
            R_22, R_27, R_32, R_37, R_42, D_22, D_27, D_32, D_37, D_42, \
            R_22_Y, R_27_Y, R_32_Y, R_37_Y, R_42_Y, D_22_Y, D_27_Y, D_32_Y, D_37_Y, D_42_Y = row
            if K is None or K_Y is None or (K <= -3 or K >= 0):
                continue
            para = (C, K, theta, C_Y, K_Y, theta_Y)
            factor = (factor_22, factor_27, factor_32, factor_37, factor_42)

            yuv_dir = file.split('.')[0] + '_' + str(poc) + '_' + str(ctu_addr) + '.yuv'
            image_dir = os.path.join(data_root, mode, yuv_dir)
            sample = (R_22, R_27, R_32, R_37, R_42, D_22, D_27, D_32, D_37, D_42)
            sample_Y = (R_22_Y, R_27_Y, R_32_Y, R_37_Y, R_42_Y, D_22_Y, D_27_Y, D_32_Y, D_37_Y, D_42_Y)
            item = (image_dir, para, sample, sample_Y)
            image_info.append(item)
    print('Image Parse Completed, total %d samples' % len(image_info))
    return image_info


class HDRDataset(torch.utils.data.Dataset):
    def __init__(
            self,
            excel_root,
            data_root,
            mode
    ):
        self.data_root = data_root
        self.excel_root = excel_root
        self.mode = mode
        self.image_info = parse_hdr(self.excel_root, self.data_root, self.mode)

    def __getitem__(self, index):
        (
            image_dir,
            para,
            sample,
            sample_Y
        ) = self.image_info[index]
        try:
            img = readYUV(image_dir)
            img = img / 1023.0
            outputs = [
                img,
                para,
                sample,
                sample_Y
            ]
        except Exception as e:
            print("problem in, ", image_dir)
            print(e)
            return self.__getitem__(np.random.randint(0, len(self.image_info)))
        return outputs

    def __len__(self):
        return len(self.image_info)


def readYUV(image_dir):
    height, width = 128, 128
    fp = open(image_dir, 'rb')
    h_h = height // 2
    h_w = width // 2
    bytes2num = partial(int.from_bytes, byteorder='little', signed=False)
    Yt = np.zeros(shape=(height, width), dtype='uint16', order='C')
    Ut = np.zeros(shape=(h_h, h_w), dtype='uint16', order='C')
    Vt = np.zeros(shape=(h_h, h_w), dtype='uint16', order='C')
    YY = np.zeros_like(Yt, dtype='uint16')
    UU = np.zeros_like(Yt, dtype='uint16')
    VV = np.zeros_like(Yt, dtype='uint16')
    YUV = np.zeros(shape=(height, width, 3), dtype='uint16')

    for m in range(height):
        for n in range(width):
            Yt[m, n] = bytes2num(fp.read(2))
    for m in range(h_h):
        for n in range(h_w):
            Ut[m, n] = bytes2num(fp.read(2))
    for m in range(h_h):
        for n in range(h_w):
            Vt[m, n] = bytes2num(fp.read(2))

    UU[0: height - 1: 2, 0: width - 1: 2] = Ut[:, :]
    UU[0: height - 1: 2, 1: width: 2] = Ut[:, :]
    UU[1: height: 2, 0: width - 1: 2] = Ut[:, :]
    UU[1: height: 2, 1: width: 2] = Ut[:, :]
    VV[0: height - 1: 2, 0: width - 1: 2] = Vt[:, :]
    VV[0: height - 1: 2, 1: width: 2] = Vt[:, :]
    VV[1: height: 2, 0: width - 1: 2] = Vt[:, :]
    VV[1: height: 2, 1: width: 2] = Vt[:, :]
    YUV[:, :, 0] = Yt
    YUV[:, :, 1] = UU
    YUV[:, :, 2] = VV
    return YUV

# from tqdm import tqdm
#
# train_data = HDRDataset('/data0/Yuan/HDRCoding/HDR_data/excel', '/data0/Yuan/HDRCoding/HDR_data/image', 'train')
# data_loader = DataLoader(train_data, batch_size=128, shuffle=True)
# iterator = tqdm(data_loader, maxinterval=100,
#                 mininterval=1, ncols=100,
#                 bar_format='{l_bar}|{bar}| {n_fmt}/{total_fmt} [{rate_fmt}{postfix}|{elapsed}<{remaining}]',
#                 smoothing=0.01)
# for i, data in enumerate(iterator):
#     pass